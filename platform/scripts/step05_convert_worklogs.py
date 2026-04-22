"""Step 05 | 工作日志 Excel → PDF。

每个 xlsx 期望含 3 个 sheet:日志正文 / 随手绘 / 工作照,
渲染为 3 页 A4 PDF,输出到 `data/output/worklog_pdfs/<YYYY-MM-DD>_worklog.pdf`。
"""
from __future__ import annotations

import datetime
import os
import re
import sys
from io import BytesIO
from pathlib import Path

import openpyxl
from PIL import Image as PILImage
from reportlab.lib.colors import HexColor, white
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from _common import get_logger, get_paths, load_config

STEP_ID = "step05"

PAGE_W, PAGE_H = A4
MARGIN = 2.0 * cm
CONTENT_W = PAGE_W - 2 * MARGIN

# 按优先级搜索 CJK 字体(Win / macOS / Linux 常见路径),找不到则回退 Helvetica。
_FONT_CANDIDATES = [
    (r"C:\Windows\Fonts\msyh.ttc", "MSYaHei"),
    (r"C:\Windows\Fonts\msyhbd.ttc", "MSYaHeiBold"),
    (r"C:\Windows\Fonts\simhei.ttf", "SimHei"),
    (r"C:\Windows\Fonts\simsun.ttc", "SimSun"),
    ("/System/Library/Fonts/PingFang.ttc", "PingFang"),
    ("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc", "NotoSansCJK"),
    ("/usr/share/fonts/truetype/arphic/uming.ttc", "UMing"),
]

_FONT_REGISTERED = False
_FONT_NAME = "Helvetica"


def register_fonts(log) -> None:
    global _FONT_REGISTERED, _FONT_NAME
    if _FONT_REGISTERED:
        return
    for fp, name in _FONT_CANDIDATES:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(name, fp))
                _FONT_NAME = name
                log.info(f"已注册中文字体: {name} ({fp})")
                break
            except Exception as e:
                log.warning(f"字体注册失败 {fp}: {e}")
    if _FONT_NAME == "Helvetica":
        log.warning("未找到可用 CJK 字体，PDF 中的中文可能显示为方框。")
    _FONT_REGISTERED = True


_EXCEL_EPOCH = datetime.datetime(1899, 12, 30)


def excel_serial_to_date(serial) -> str:
    try:
        n = int(serial)
        dt = _EXCEL_EPOCH + datetime.timedelta(days=n)
        return f"{dt.year}年{dt.month}月{dt.day}日"
    except (ValueError, TypeError, OverflowError):
        return str(serial)


def format_date_value(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        return f"{val.year}年{val.month}月{val.day}日"
    if isinstance(val, datetime.date):
        return f"{val.year}年{val.month}月{val.day}日"
    s = str(val).strip().replace('\t', '')
    if not s:
        return ""
    if s.isdigit() and len(s) >= 5:
        return excel_serial_to_date(int(s))
    return s


def extract_images_from_sheet(ws) -> list[PILImage.Image]:
    images = []
    if not hasattr(ws, '_images'):
        return images
    for img in ws._images:
        try:
            data = BytesIO(img._data())
            pil = PILImage.open(data)
            pil.load()
            if pil.mode == 'RGBA':
                pil = pil.convert('RGB')
            images.append(pil)
        except Exception:
            pass
    return images


def get_unique_body_texts(ws, start_row: int) -> list[str]:
    merged = list(ws.merged_cells.ranges)
    visited = set()
    texts = []
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        in_merged = False
        for mr in merged:
            if (mr.min_row <= row_idx <= mr.max_row
                    and mr.min_col <= 1 <= mr.max_col):
                in_merged = True
                key = str(mr)
                if key in visited:
                    break
                visited.add(key)
                val = ws.cell(row=mr.min_row, column=mr.min_col).value
                if val is not None and str(val).strip():
                    texts.append(str(val).strip())
                break
        if not in_merged and cell.value is not None and str(cell.value).strip():
            texts.append(str(cell.value).strip())
    return texts


def draw_page_frame(c, title: str, date_str="", weather_str="", recorder=""):
    c.setStrokeColor(HexColor("#3a5a8c"))
    c.setLineWidth(1.5)
    c.roundRect(MARGIN - 10, MARGIN - 10,
                CONTENT_W + 20, PAGE_H - 2 * MARGIN + 20, 6)

    c.setFillColor(HexColor("#1a3a5c"))
    c.rect(MARGIN - 10, PAGE_H - MARGIN - 36, CONTENT_W + 20, 40, fill=True, stroke=False)

    c.setFillColor(white)
    c.setFont(_FONT_NAME, 14)
    c.drawCentredString(PAGE_W / 2, PAGE_H - MARGIN - 24, title)

    y = PAGE_H - MARGIN - 50
    if date_str or weather_str:
        y -= 8
        c.setFont(_FONT_NAME, 10)
        c.setFillColor(HexColor("#333333"))
        if date_str:
            c.drawString(MARGIN, y, f"日期：{date_str}")
        if weather_str:
            c.drawRightString(PAGE_W - MARGIN, y, f"天气：{weather_str}")
        y -= 6
        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.5)
        c.line(MARGIN, y, PAGE_W - MARGIN, y)
        y -= 14
    if recorder:
        c.setFont(_FONT_NAME, 10)
        c.setFillColor(HexColor("#555555"))
        c.drawString(MARGIN, y, f"记录人：{recorder}")
        y -= 18
    if not date_str and not weather_str and not recorder:
        y -= 14
    return y


def wrap_text_lines(text: str, font_name: str, font_size: float, max_width: float) -> list[str]:
    lines = []
    for raw in text.split('\n'):
        raw = raw.rstrip()
        if not raw:
            lines.append("")
            continue
        current = ""
        for ch in raw:
            test = current + ch
            if pdfmetrics.stringWidth(test, font_name, font_size) > max_width:
                lines.append(current)
                current = ch
            else:
                current = test
        if current:
            lines.append(current)
    return lines


def render_sheet_log(c, ws, default_title: str, date_from_filename=""):
    raw_title = ws.cell(row=1, column=1).value
    title = str(raw_title).strip() if raw_title else default_title

    date_cell = ws.cell(row=3, column=2).value
    date_str = format_date_value(date_cell) or date_from_filename

    weather_cell = ws.cell(row=3, column=9).value
    if weather_cell is None:
        for col in range(8, ws.max_column + 1):
            v = ws.cell(row=3, column=col).value
            if v is not None and str(v).strip() and str(v).strip() != "天气：":
                weather_cell = v
                break
    weather_str = str(weather_cell).strip() if weather_cell else ""

    recorder_cell = ws.cell(row=4, column=2).value
    recorder = str(recorder_cell).strip() if recorder_cell else ""

    y = draw_page_frame(c, title, date_str, weather_str, recorder)
    body = "\n".join(get_unique_body_texts(ws, 5))
    if body.strip():
        font_size = 9.5
        line_height = 16
        lines = wrap_text_lines(body.strip(), _FONT_NAME, font_size, CONTENT_W - 10)
        c.setFont(_FONT_NAME, font_size)
        c.setFillColor(HexColor("#222222"))
        for line in lines:
            if y < MARGIN + 20:
                break
            c.drawString(MARGIN + 5, y, line)
            y -= line_height


def render_sheet_sketch(c, ws):
    y = draw_page_frame(c, "工作日志 · 随手绘")
    drawer_cell = ws.cell(row=4, column=2).value
    if drawer_cell and str(drawer_cell).strip():
        c.setFont(_FONT_NAME, 10)
        c.setFillColor(HexColor("#555555"))
        c.drawString(MARGIN, y, f"绘图：{str(drawer_cell).strip()}")
        y -= 20

    images = extract_images_from_sheet(ws)
    if images:
        for img in images:
            if y < MARGIN + 40:
                break
            iw, ih = img.size
            scale = min(CONTENT_W / iw, (y - MARGIN - 10) / ih, 1.0)
            dw, dh = iw * scale, ih * scale
            x = MARGIN + (CONTENT_W - dw) / 2
            c.drawImage(ImageReader(img), x, y - dh, dw, dh, preserveAspectRatio=True)
            y -= dh + 10
    else:
        c.setFont(_FONT_NAME, 11)
        c.setFillColor(HexColor("#999999"))
        c.drawCentredString(PAGE_W / 2, PAGE_H / 2, "（本页无随手绘内容）")


def render_sheet_photo(c, ws):
    y = draw_page_frame(c, "工作日志 · 工作照")
    captions = get_unique_body_texts(ws, 4)
    images = extract_images_from_sheet(ws)

    if images:
        n = len(images)
        if n == 1:
            img = images[0]
            iw, ih = img.size
            max_h = y - MARGIN - 40
            scale = min(CONTENT_W / iw, max_h / ih, 1.0)
            dw, dh = iw * scale, ih * scale
            x = MARGIN + (CONTENT_W - dw) / 2
            c.drawImage(ImageReader(img), x, y - dh, dw, dh, preserveAspectRatio=True)
            y -= dh + 5
            if captions:
                c.setFont(_FONT_NAME, 8)
                c.setFillColor(HexColor("#666666"))
                c.drawCentredString(PAGE_W / 2, y, captions[0])
        else:
            cols = 2
            gap = 8
            img_w = (CONTENT_W - gap) / cols
            max_img_h = img_w * 0.75
            for i, img in enumerate(images):
                if y < MARGIN + 30:
                    break
                col = i % cols
                if col == 0 and i > 0:
                    y -= max_img_h + 24
                if y < MARGIN + 30:
                    break
                x = MARGIN + col * (img_w + gap)
                iw, ih = img.size
                scale = min(img_w / iw, max_img_h / ih, 1.0)
                dw, dh = iw * scale, ih * scale
                ix = x + (img_w - dw) / 2
                c.drawImage(ImageReader(img), ix, y - dh, dw, dh, preserveAspectRatio=True)
                if i < len(captions):
                    c.setFont(_FONT_NAME, 7)
                    c.setFillColor(HexColor("#666666"))
                    c.drawCentredString(x + img_w / 2, y - dh - 10, captions[i][:30])
            y -= max_img_h + 24
    elif captions:
        c.setFont(_FONT_NAME, 10)
        c.setFillColor(HexColor("#333333"))
        for cap in captions:
            if y < MARGIN + 20:
                break
            c.drawString(MARGIN + 5, y, cap)
            y -= 16
    else:
        c.setFont(_FONT_NAME, 11)
        c.setFillColor(HexColor("#999999"))
        c.drawCentredString(PAGE_W / 2, PAGE_H / 2, "（本页无工作照内容）")


def convert_one(xlsx_path: Path, pdf_path: Path, default_title: str, date_from_filename=""):
    wb = openpyxl.load_workbook(str(xlsx_path))
    sheets = wb.sheetnames

    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.setTitle(xlsx_path.stem)

    if len(sheets) >= 1:
        render_sheet_log(c, wb[sheets[0]], default_title, date_from_filename)
        c.showPage()
    if len(sheets) >= 2:
        render_sheet_sketch(c, wb[sheets[1]])
        c.showPage()
    if len(sheets) >= 3:
        render_sheet_photo(c, wb[sheets[2]])
        c.showPage()

    c.save()
    wb.close()


def extract_date_from_filename(name: str) -> str:
    m = re.match(r'(\d{4})(\d{2})(\d{2})', name)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y}年{mo}月{d}日"
    return ""


def extract_date_key_from_filename(name: str) -> str:
    m = re.match(r'(\d{4})(\d{2})(\d{2})', name)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return ""


def main() -> int:
    log = get_logger(STEP_ID)
    cfg = load_config()
    paths = get_paths()

    register_fonts(log)

    project = cfg.get("project") or {}
    default_title = project.get("full_name") or (
        (project.get("name") or "") + "第四次全国文物普查工作日志"
    ).strip() or "第四次全国文物普查工作日志"

    in_dir = paths.input_worklogs
    out_dir = paths.output_worklogs

    log.info("=" * 70)
    log.info("Step 05 | 工作日志 Excel → PDF")
    log.info(f"  输入: {in_dir}")
    log.info(f"  输出: {out_dir}")
    log.info(f"  页眉默认标题: {default_title}")
    log.info("=" * 70)

    if not in_dir.exists():
        log.error(f"未找到工作日志输入: {in_dir}")
        return 11

    out_dir.mkdir(parents=True, exist_ok=True)

    files = sorted(in_dir.glob("*.xlsx"))
    # 过滤 Office 临时锁文件。
    files = [f for f in files if not f.name.startswith("~$")]

    if not files:
        log.warning("未发现任何 .xlsx 文件。")
        return 0

    log.info(f"发现 {len(files)} 个 Excel 文件")
    success = 0
    failed = 0
    for i, f in enumerate(files, 1):
        date_key = extract_date_key_from_filename(f.stem)
        date_readable = extract_date_from_filename(f.stem)
        pdf_name = f"{date_key}_worklog.pdf" if date_key else f"{f.stem}.pdf"
        pdf_path = out_dir / pdf_name
        try:
            convert_one(f, pdf_path, default_title, date_readable)
            success += 1
            log.info(f"  [{i}/{len(files)}] OK  {f.name} → {pdf_name}")
        except Exception as e:
            failed += 1
            log.error(f"  [{i}/{len(files)}] FAIL {f.name}: {e}")

    log.info("-" * 70)
    log.info(f"转换完成: 成功 {success}, 失败 {failed}")
    log.info(f"PDF 输出目录: {out_dir}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
