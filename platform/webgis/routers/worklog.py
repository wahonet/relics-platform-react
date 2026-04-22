"""工作日志 API。

合并 step05 产出的每日 PDF (`data/output/worklog_pdfs/YYYY-MM-DD_*.pdf`)
与总台账 Excel (`data/input/02_worklogs/*.xlsx`,文件名含"台账"或"ledger")。
"""
from __future__ import annotations

import datetime as _dt
import re
import sys
from pathlib import Path

from fastapi import APIRouter

_SCRIPTS_DIR = Path(__file__).resolve().parents[2] / "scripts"
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))

from _common import get_paths  # noqa: E402

router = APIRouter(tags=["工作日志"])

_EXCEL_EPOCH = _dt.datetime(1899, 12, 30)

_ledger_cache: list[dict] | None = None
_ledger_by_date: dict[str, dict] | None = None


def _parse_date(val) -> str:
    if isinstance(val, _dt.datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, _dt.date):
        return val.strftime("%Y-%m-%d")
    if val is None:
        return ""
    s = str(val).strip()
    if s.isdigit() and len(s) >= 5:
        dt = _EXCEL_EPOCH + _dt.timedelta(days=int(s))
        return dt.strftime("%Y-%m-%d")
    return s


def _find_ledger_path() -> Path | None:
    """02_worklogs/ 下文件名含"台账"或"ledger"的第一个 xlsx。"""
    worklog_dir = get_paths().input_worklogs
    if not worklog_dir.exists():
        return None
    for f in sorted(worklog_dir.glob("*.xlsx")):
        name = f.stem.lower()
        if "台账" in f.stem or "ledger" in name:
            return f
    return None


def _load_ledger() -> None:
    """读取并缓存台账。列顺序约定:日期 / 时长 / 人数 / 参与人 / 镇街 / 村 /
    复核数 / 复核名 / 新发现数 / 新发现名 / 消亡数 / 消亡名 / 备注,从第 3 行起。"""
    global _ledger_cache, _ledger_by_date
    if _ledger_cache is not None:
        return

    _ledger_cache = []
    _ledger_by_date = {}

    path = _find_ledger_path()
    if not path:
        return

    try:
        import openpyxl
    except ImportError:
        return

    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        vals = [c.value for c in row]
        if not vals or vals[0] is None:
            continue
        date_str = _parse_date(vals[0])
        if not date_str:
            continue

        def _at(i: int, default=""):
            if i < len(vals) and vals[i] is not None:
                return vals[i]
            return default

        record = {
            "date": date_str,
            "duration": _at(1, ""),
            "participant_count": _at(2, ""),
            "participants": str(_at(3, "")).strip(),
            "township": str(_at(4, "")).strip(),
            "villages": str(_at(5, "")).strip(),
            "review_count": _at(6, 0),
            "review_names": str(_at(7, "")).strip(),
            "new_count": _at(8, 0),
            "new_names": str(_at(9, "")).strip(),
            "lost_count": _at(10, 0),
            "lost_names": str(_at(11, "")).strip(),
            "remark": str(_at(12, "")).strip(),
        }
        _ledger_cache.append(record)
        _ledger_by_date[date_str] = record

    wb.close()


def _get_pdf_list() -> dict[str, str]:
    """扫描 worklog_pdfs 目录,按文件名前缀 YYYY-MM-DD 建立索引。"""
    pdf_dir = get_paths().output_worklogs
    result: dict[str, str] = {}
    if not pdf_dir.exists():
        return result
    for f in pdf_dir.glob("*.pdf"):
        m = re.match(r"(\d{4}-\d{2}-\d{2})", f.name)
        if m:
            result[m.group(1)] = f.name
    return result


@router.get("/worklog/dates")
async def worklog_dates():
    _load_ledger()
    pdf_map = _get_pdf_list()

    all_dates = sorted(set(list(pdf_map.keys()) + list((_ledger_by_date or {}).keys())))
    items = []
    for d in all_dates:
        ledger = (_ledger_by_date or {}).get(d, {})
        items.append({
            "date": d,
            "has_pdf": d in pdf_map,
            "pdf_file": pdf_map.get(d, ""),
            "township": ledger.get("township", ""),
            "villages": ledger.get("villages", ""),
            "participants": ledger.get("participants", ""),
            "review_count": ledger.get("review_count", 0),
            "review_names": ledger.get("review_names", ""),
            "new_count": ledger.get("new_count", 0),
            "new_names": ledger.get("new_names", ""),
            "lost_count": ledger.get("lost_count", 0),
            "lost_names": ledger.get("lost_names", ""),
        })

    return {"total_days": len(all_dates), "items": items}


@router.get("/worklog/detail/{date}")
async def worklog_detail(date: str):
    _load_ledger()
    pdf_map = _get_pdf_list()

    ledger = (_ledger_by_date or {}).get(date, {})
    return {
        "date": date,
        "has_pdf": date in pdf_map,
        "pdf_file": pdf_map.get(date, ""),
        "ledger": ledger if ledger else None,
    }
